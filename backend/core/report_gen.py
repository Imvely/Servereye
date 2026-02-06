"""엑셀 리포트 생성 모듈"""
import logging
import os
from datetime import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from backend.db.database import async_session
from backend.config import REPORTS_DIR

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


async def generate_report(date_from: str, date_to: str,
                          server_ids: Optional[list[int]] = None,
                          report_type: str = 'summary',
                          report_name: Optional[str] = None,
                          created_by: str = 'system') -> dict:
    """리포트 생성"""
    wb = Workbook()

    # 요약 시트
    await _create_summary_sheet(wb, date_from, date_to, server_ids)

    # 서버별 현황 시트
    await _create_server_sheet(wb, date_from, date_to, server_ids)

    if report_type in ('detail', 'summary'):
        # 시계열 데이터 시트
        await _create_timeseries_sheet(wb, date_from, date_to, server_ids)

    # 알림 이력 시트
    await _create_alert_sheet(wb, date_from, date_to, server_ids)

    # 차트 시트
    await _create_chart_sheet(wb, date_from, date_to, server_ids)

    # 파일 저장
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{report_name or 'ServerEye_Report'}_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)

    file_size_kb = os.path.getsize(filepath) // 1024

    # 리포트 이력 저장
    async with async_session() as session:
        await session.execute(
            text("""INSERT INTO report_history
                (report_name, report_type, server_ids, date_from, date_to,
                 file_path, file_size_kb, created_by)
                VALUES (:rn, :rt, :si, :df, :dt, :fp, :fs, :cb)"""),
            {
                "rn": filename, "rt": report_type,
                "si": str(server_ids) if server_ids else None,
                "df": date_from, "dt": date_to,
                "fp": filepath, "fs": file_size_kb, "cb": created_by
            }
        )
        await session.commit()
        result = await session.execute(text("SELECT last_insert_rowid()"))
        report_id = result.scalar()

    return {
        "report_id": report_id,
        "filename": filename,
        "filepath": filepath,
        "file_size_kb": file_size_kb
    }


def _set_header(ws, row, columns):
    """헤더 스타일 설정"""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


async def _create_summary_sheet(wb, date_from, date_to, server_ids):
    """요약 시트"""
    ws = wb.active
    ws.title = "요약"

    ws['A1'] = 'ServerEye 모니터링 리포트'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A3'] = f'기간: {date_from} ~ {date_to}'
    ws['A4'] = f'생성일: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    async with async_session() as session:
        # 서버 수
        sid_filter = ""
        params = {"df": date_from, "dt": date_to}
        if server_ids:
            placeholders = ','.join(str(s) for s in server_ids)
            sid_filter = f"AND server_id IN ({placeholders})"

        result = await session.execute(
            text(f"SELECT COUNT(*) FROM servers WHERE is_active=1 {sid_filter}")
        )
        total = result.scalar()
        ws['A6'] = f'대상 서버: {total}대'

        # 가동률
        result = await session.execute(
            text(f"""SELECT
                ROUND(100.0 * SUM(CASE WHEN status != 'offline' THEN 1 ELSE 0 END) / COUNT(*), 1)
                FROM servers WHERE is_active=1 {sid_filter}""")
        )
        uptime = result.scalar() or 0
        ws['A7'] = f'평균 가동률: {uptime}%'

        # 알림 건수
        result = await session.execute(
            text(f"""SELECT COUNT(*) FROM alert_history
                 WHERE created_at BETWEEN :df AND :dt {sid_filter}"""),
            params
        )
        alert_count = result.scalar()
        ws['A8'] = f'기간 내 알림 발생: {alert_count}건'

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25


async def _create_server_sheet(wb, date_from, date_to, server_ids):
    """서버별 현황 시트"""
    ws = wb.create_sheet("서버별 현황")
    columns = ['서버명', 'IP', 'OS', '그룹', '상태', '평균CPU%', '최대CPU%',
               '평균MEM%', '최대MEM%', '가동률%', '알림수']
    _set_header(ws, 1, columns)

    sid_filter = ""
    if server_ids:
        placeholders = ','.join(str(s) for s in server_ids)
        sid_filter = f"AND s.server_id IN ({placeholders})"

    async with async_session() as session:
        result = await session.execute(
            text(f"""SELECT s.display_name, s.ip_address, s.os_type, s.group_name, s.status,
                ROUND(AVG(m.cpu_avg), 1), ROUND(MAX(m.cpu_max), 1),
                ROUND(AVG(m.mem_avg_pct), 1), ROUND(MAX(m.mem_max_pct), 1),
                (SELECT COUNT(*) FROM alert_history a
                 WHERE a.server_id=s.server_id AND a.created_at BETWEEN :df AND :dt)
                FROM servers s
                LEFT JOIN metrics_hourly m ON s.server_id=m.server_id
                    AND m.bucket_time BETWEEN :df AND :dt
                WHERE s.is_active=1 {sid_filter}
                GROUP BY s.server_id"""),
            {"df": date_from, "dt": date_to}
        )
        rows = result.fetchall()

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


async def _create_timeseries_sheet(wb, date_from, date_to, server_ids):
    """시계열 데이터 시트"""
    ws = wb.create_sheet("시계열 데이터")
    columns = ['시간', '서버', 'CPU%', 'MEM%', 'DISK_READ', 'DISK_WRITE']
    _set_header(ws, 1, columns)

    sid_filter = ""
    if server_ids:
        placeholders = ','.join(str(s) for s in server_ids)
        sid_filter = f"AND m.server_id IN ({placeholders})"

    async with async_session() as session:
        result = await session.execute(
            text(f"""SELECT m.bucket_time, s.display_name,
                m.cpu_avg, m.mem_avg_pct, m.disk_read_avg, m.disk_write_avg
                FROM metrics_hourly m
                JOIN servers s ON m.server_id=s.server_id
                WHERE m.bucket_time BETWEEN :df AND :dt {sid_filter}
                ORDER BY m.bucket_time"""),
            {"df": date_from, "dt": date_to}
        )
        rows = result.fetchall()

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


async def _create_alert_sheet(wb, date_from, date_to, server_ids):
    """알림 이력 시트"""
    ws = wb.create_sheet("알림 이력")
    columns = ['발생시각', '서버', '심각도', '메트릭', '값', '임계치', '해결시각', '소요시간']
    _set_header(ws, 1, columns)

    sid_filter = ""
    if server_ids:
        placeholders = ','.join(str(s) for s in server_ids)
        sid_filter = f"AND a.server_id IN ({placeholders})"

    async with async_session() as session:
        result = await session.execute(
            text(f"""SELECT a.created_at, s.display_name, a.severity,
                a.metric_name, a.metric_value, a.threshold_value,
                a.resolved_at,
                CASE WHEN a.resolved_at IS NOT NULL
                    THEN ROUND((julianday(a.resolved_at) - julianday(a.created_at)) * 86400)
                    ELSE NULL END
                FROM alert_history a
                JOIN servers s ON a.server_id=s.server_id
                WHERE a.created_at BETWEEN :df AND :dt {sid_filter}
                ORDER BY a.created_at DESC"""),
            {"df": date_from, "dt": date_to}
        )
        rows = result.fetchall()

    severity_fills = {
        'critical': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        'warning': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    }

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
        severity = row[2] if len(row) > 2 else ''
        if severity in severity_fills:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = severity_fills[severity]

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


async def _create_chart_sheet(wb, date_from, date_to, server_ids):
    """차트 시트"""
    ws = wb.create_sheet("차트")
    ws['A1'] = 'CPU/MEM 추세 차트'
    ws['A1'].font = Font(bold=True, size=14)

    sid_filter = ""
    if server_ids:
        placeholders = ','.join(str(s) for s in server_ids)
        sid_filter = f"AND server_id IN ({placeholders})"

    async with async_session() as session:
        result = await session.execute(
            text(f"""SELECT bucket_time,
                ROUND(AVG(cpu_avg), 1), ROUND(AVG(mem_avg_pct), 1)
                FROM metrics_hourly
                WHERE bucket_time BETWEEN :df AND :dt {sid_filter}
                GROUP BY bucket_time ORDER BY bucket_time"""),
            {"df": date_from, "dt": date_to}
        )
        rows = result.fetchall()

    if not rows:
        ws['A3'] = '데이터 없음'
        return

    ws['A3'] = '시간'
    ws['B3'] = 'CPU 평균%'
    ws['C3'] = 'MEM 평균%'

    for i, row in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3, value=row[2])

    if len(rows) > 1:
        chart = LineChart()
        chart.title = "CPU/MEM 추세"
        chart.y_axis.title = "%"
        chart.x_axis.title = "시간"
        chart.width = 30
        chart.height = 15

        data_end = 3 + len(rows)
        cpu_data = Reference(ws, min_col=2, min_row=3, max_row=data_end)
        mem_data = Reference(ws, min_col=3, min_row=3, max_row=data_end)
        dates = Reference(ws, min_col=1, min_row=4, max_row=data_end)

        chart.add_data(cpu_data, titles_from_data=True)
        chart.add_data(mem_data, titles_from_data=True)
        chart.set_categories(dates)

        ws.add_chart(chart, "A" + str(data_end + 2))
